from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from io import BytesIO, StringIO
import csv

def export_orders_to_excel(orders_data: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    
    headers = ["ID", "Пользователь", "Дата", "Статус", "Сумма", "Блюда"]
    ws.append(headers)
    
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for order in orders_data:
        dishes_text = ", ".join([f"{item['name']} x{item['quantity']}" for item in order.get('items', [])])
        ws.append([
            order['id'],
            order['user_name'],
            order['date'].strftime('%d.%m.%Y'),
            order['status'],
            order['total_amount'],
            dishes_text
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_statistics_to_excel(stats_data: dict) -> BytesIO:
    wb = Workbook()
    
    if 'summary' in stats_data:
        ws = wb.active
        ws.title = "Сводка"
        ws.append(["Показатель", "Значение"])
        ws.append(["Всего заказов", stats_data['summary']['total_orders']])
        ws.append(["Уникальных пользователей", stats_data['summary']['unique_users']])
        ws.append(["Общая сумма", stats_data['summary']['total_amount']])
    
    if 'dishes' in stats_data:
        ws = wb.create_sheet("Блюда")
        ws.append(["Блюдо", "Количество", "Выручка"])
        for dish in stats_data['dishes']:
            ws.append([dish['name'], dish['quantity'], dish['revenue']])
    
    if 'users' in stats_data:
        ws = wb.create_sheet("Пользователи")
        ws.append(["Пользователь", "Заказов", "Сумма", "Средний чек"])
        for user in stats_data['users']:
            ws.append([
                user['name'],
                user['orders_count'],
                user['total_amount'],
                user['avg_order']
            ])
    
    for ws in wb.worksheets:
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_statistics_to_csv(stats_data: dict) -> BytesIO:
    output = StringIO()
    writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    
    summary = stats_data.get("summary", {})
    writer.writerow(["Сводка заказов"])
    writer.writerow(["Всего заказов", summary.get("total_orders", 0)])
    writer.writerow(["Уникальных пользователей", summary.get("unique_users", 0)])
    writer.writerow(["Общая сумма", f"{summary.get('total_amount', 0):.2f}"])
    writer.writerow([])
    
    dishes = stats_data.get("dishes", [])
    if dishes:
        writer.writerow(["Статистика по блюдам"])
        writer.writerow(["Название", "Количество", "Выручка"])
        for dish in dishes:
            writer.writerow([
                dish.get("name", ""),
                dish.get("quantity", 0),
                f"{dish.get('revenue', 0):.2f}"
            ])
        writer.writerow([])
    
    users = stats_data.get("users", [])
    if users:
        writer.writerow(["Статистика по пользователям"])
        writer.writerow(["Имя", "Заказов", "Сумма", "Средний чек"])
        for user in users:
            writer.writerow([
                user.get("name", ""),
                user.get("orders_count", 0),
                f"{user.get('total_amount', 0):.2f}",
                f"{user.get('avg_order', 0):.2f}"
            ])
    
    csv_content = output.getvalue()
    csv_file = BytesIO()
    csv_file.write(csv_content.encode('utf-8-sig'))
    csv_file.seek(0)
    return csv_file

def export_cafe_report_to_excel(cafe_report: dict) -> BytesIO:
    wb = Workbook()
    
    date_str = cafe_report['date'].strftime('%d.%m.%Y')
    
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    ws_summary.append(["Отчет по кафе", f"Дата: {date_str}"])
    ws_summary.append([])
    ws_summary.append(["Всего заказов", cafe_report['total_orders']])
    ws_summary.append(["Общая сумма", f"{cafe_report['total_amount']:.2f} ₽"])
    ws_summary.append(["Всего позиций", cafe_report['total_items']])
    ws_summary.append([])
    
    for cafe_data in cafe_report['cafes']:
        ws = wb.create_sheet(cafe_data['cafe_name'])
        ws.append(["Отчет по кафе", cafe_data['cafe_name']])
        ws.append(["Дата", date_str])
        ws.append([])
        ws.append(["Всего заказов", cafe_data['total_orders']])
        ws.append(["Сотрудников", cafe_data['unique_users']])
        ws.append(["Общая сумма", f"{cafe_data['total_amount']:.2f} ₽"])
        ws.append(["Всего позиций", cafe_data['total_items']])
        ws.append([])
        ws.append(["Сотрудник", "Блюда", "Время доставки", "Тип доставки", "Сумма"])
        
        for order in cafe_data['orders']:
            ws.append([
                order['user_name'],
                order['items'],
                order.get('delivery_time', ''),
                order.get('delivery_type', ''),
                f"{order['total']:.2f} ₽"
            ])
    
    for ws in wb.worksheets:
        header_font = Font(bold=True)
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
        if ws.title != "Сводка":
            for row in ws.iter_rows(min_row=9, max_row=9):
                for cell in row:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output